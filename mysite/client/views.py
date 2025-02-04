import openpyxl
from openpyxl import load_workbook
from xlrd import open_workbook

from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from .forms import ClientForm, ClientVoucherForm, UploadExcelForm
from .models import Client, ClientVoucher
from service.models import Voucher


def add_client(request):
    """
    Add a new client to the database.

    This view handles the POST request to add a new client by verifying that the required fields (first name, last name)
    are provided and checks for duplicate telephone numbers or emails before creating a new client.

    :param request: HttpRequest object containing the details of the request.

    :return:
        - If POST request is successful, redirects to the 'client' page with a success message.
        - If duplicate telephone number or email is found, displays a warning message.
        - If no POST request, renders the 'add.html' template with the form.
    """
    # Gets the form data
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "El cliente fue creado correctamente.")
            return redirect(reverse('client:client_list'))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ClientForm()
        
    return render(request, 'client/add.html', {'form': form})

    
def delete_client(request, client_id):
    """
    Delete an existing client from the database.

    This view displays a confirmation page to delete a specific client. If confirmed via a POST request, it removes the
    client from the database.

    :param request: HttpRequest object with the request details.

    :param client_id: Integer ID of the client to be deleted.

    :return:
        - If POST request is successful, redirects to the 'client' page with a success message.
        - If no POST request, renders the 'client_delete.html' template to confirm deletion.
    """
    # Get the specific client or return 404 if it doesn't exist
    client = get_object_or_404(Client, id=client_id)
    # Gets the form data
    if request.method == 'POST':
        client.delete()
        messages.success(request, "Cliente eliminado correctamente.")
        return redirect('client:client_list')
    
    return render(request, 'client/delete.html', {'client': client})


def edit_client(request, client_id):
    """
    Edits the details of an existing client.

    Allows updating client information by providing new data in a POST request.

    :param request: HttpRequest object with the request details.

    :param client_id: Integer ID of the client to edit.

    :return:
        - If POST request is successful, redirects to the 'client' page.
        - If no POST request, renders the 'client_edit.html' template with current client data.
    """
    # Get the specific client or return 404 if it doesn't exist
    client = get_object_or_404(Client, id=client_id)
    
    # Gets the form data
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect(reverse('client:client_list'))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ClientForm(instance=client)
    
    return render(
        request, 'client/edit.html', {'form': form, 'client': client}
    )


def export_clients(request):
    """
    Exports the list of clients to an Excel (.xlsx) file.

    This function creates an Excel workbook with client data, including firs name, last name, telephone number and email
     of each client. It sends the files as a response for download.

    :return: HttpResponse object with the Excel file content a headers for download.
    """
    # Create a new work book
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    # Write column headers
    ws.append(['Nombre', 'Apellidos', 'Teléfono', 'email'])
    # Get all clients
    clients = Client.objects.all()
    # Write clients data
    for client in clients:
        ws.append([client.first_name, client.last_name, client.telephone_number, client.email])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=listado_clientes.xlsx'
    wb.save(response)
    
    return response


def import_clients(request):
    """
    Handles the import of client data from an Excel file (.xls or .xlsx).

    Allows clients to upload an Excel file, validates the contents, checks for duplicates, and adds valid data to the
    database.

    :param request: HttpRequest object with the request details.

    :return:
        - If POST request and form is valid, redirects to the 'client' page.
        - If POST request and form is invalid, displays error messages.
        - If GET request, renders the 'import.html' template with the form.
    """
    if request.method == 'POST':
        # Allow the client to upload an Excel file.
        form = UploadExcelForm(request.POST, request.FILES)
        
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            successful_imports = 0
            failed_imports = 0
            
            try:
                if excel_file.name.lower().endswith('.xlsx'):
                    wb = load_workbook(excel_file)
                    sheet = wb.active
                    rows = sheet.iter_rows(min_row=2, values_only=True)
                elif excel_file.name.lower().endswith('.xls'):
                    wb = open_workbook(file_contents=excel_file.read())
                    sheet = wb.sheet_by_index(0)
                    rows = (sheet.row_values(row_idx) for row_idx in range(1, sheet.nrows))
                else:
                    messages.error(request, "Formato de archivo no válido.")
                    return render(request, 'client/import.html', {'form': form})
                
                # Process rows from file
                for row in rows:
                    try:
                        first_name, last_name, telephone_number, email = row[:4]
                        Client.objects.create(
                            first_name=first_name,
                            last_name=last_name,
                            telephone_number=telephone_number,
                            email=email
                        )
                        successful_imports += 1
                    except Exception:
                        failed_imports += 1
              
                messages.success(
                    request,
            f"Importación completada: {successful_imports} clientes importados exitosamente."
                    f"{failed_imports} fallidos."
                )
                return redirect(reverse('client:client_list'))
            
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return render(request, 'client/import.html', {'form': form})
        else:
            
            messages.error(request, "Por favor, sube un archivo válido.")
            return render(request, 'client/import.html', {'form': form})
        
    # In case of GET or if the form is invalid, redisplay the form.
    form = UploadExcelForm()
    return render(request, 'client/import.html', {'form': form})


def list_client(request):
    """
    Retrieves and display all clients with pagination.

    This view handles the request to the URL '/list/' and renders a template with a paginated list of clients, ordered
    by their first name. Handles invalid page number and ensuring the client is shown the first or last page if the
    request page is out of range.

    :param request: HttpRequest object, which may include the optional 'page' GET parameter for pagination.

    :return: HttpResponse response with the rendered 'client_list.html' template with the paginated list of clients.

    :raises:
        - PageNotAnInteger: If the page parameter is not an integer, the first page of results is shown.
        - EmpyPage: If the page parameter exceeds the total number of pages, the last page is shown.
    """
    clients = Client.objects.all().order_by('first_name')
    # Pagination with 20 clients per page
    paginator = Paginator(clients, 20)
    page_number = request.GET.get('page')
    try:
        clients = paginator.page(page_number)
    except PageNotAnInteger:
        # If page_number is not an integer get the first page
        clients = paginator.page(1)
    except EmptyPage:
        # If page_number is out of range get last page of results
        clients = paginator.page(paginator.num_pages)
    
    return render(request, 'client/list.html', {'clients': clients})


def search_client(request):
    """
    Searches for clients based on the provided query.

    Searches for clients by first name, last name, or telephone number. If the query contains a space, it splits
    the query to search by first name and last name separately.

    :param request: HttpRequest object with the search query parameter.

    :return: HttpResponse with the rendered 'search_items.html' template, displaying matching clients.
    """
    query = request.GET.get('query', '').strip()
    clients = []
    
    if query:
        
        if " " in query:
            first_name_query, last_name_query = query.split(" ", 1)
            # Filter clients by first name and last name without accents
            clients = Client.objects.filter(
                Q(first_name__icontains=first_name_query) &
                Q(last_name__icontains=last_name_query)
            )
        else:
            clients = Client.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(telephone_number__icontains=query)
            )
    
    return render(request, 'client/search.html', {'clients': clients, 'query': query})


def view_detail_client(request, client_id):
    """
    Displays detailed information of a specific client.

    Retrieves a client's details and renders a template to show the information.

    :param request: HttpRequest object with the request details.
    :param client_id: Integer ID of the client to view.

    :return: HttpResponse with the rendered 'view.html' template showing client details.
    """
    client = get_object_or_404(Client, id=client_id)
    return render(request, 'client/view.html', {'client': client})


# Client voucher
def assign_voucher(request, client_id):
    """
    Assign a voucher to a specific client.

    This view allows assigning a voucher to a client by creating a ClientVoucher object linked to the client.

    :param request: HttpRequest object with the request details.

    :param client_id: Integer ID of the client to assign the voucher.

    :return: Redirects to the voucher list page for the client if successful.
    """
    client = get_object_or_404(Client, id=client_id)
    vouchers = Voucher.objects.all()
    
    if request.method == 'POST':
        form = ClientVoucherForm(request.POST)
        if form.is_valid():
            voucher = form.save(commit=False)
            voucher.client = client
            voucher.save()
            return redirect(reverse('client:voucher_client_list', args=[client.id]))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ClientVoucherForm()
        
    return render(
        request,
        'voucher/assign.html',
        {'form': form, 'client': client, 'vouchers': vouchers})

    
def delete_client_voucher(request, voucher_id):
    """
    Delete a voucher associated with a client.

    This view allows deleting a voucher from a client. It confirms the deletion via a POST request.

    :param request: HttpRequest object with the request details.

   :param voucher_id: Integer ID of the voucher to delete.

   :return: Redirects to the voucher list page for the client if successful.
   """
    voucher = get_object_or_404(ClientVoucher, id=voucher_id)
    client = voucher.client

    if request.method == 'POST':
        voucher.delete()
        
        messages.success(request, "Bono eliminado correctamente.")
        return redirect(reverse('client:voucher_client_list', args=[client.id]))
    
    return render(request, 'voucher/client_delete.html', {'voucher': voucher, 'client': client})
    

def edit_client_voucher(request, voucher_id):
    """
    Edit the details of a voucher associated with a client.

    This view allows updating the voucher details for a client.

    :param request: HttpRequest object with the request details.
    
    :param voucher_id: Integer ID of the voucher to edit.
    
    :return: Redirects to the voucher list page for the client if successful, or re-renders the form with errors if
             invalid.
    """
    voucher = get_object_or_404(ClientVoucher, id=voucher_id)
    client = voucher.client
    
    # Gets the form data
    if request.method == 'POST':
        form = ClientVoucherForm(request.POST, instance=voucher)
        if form.is_valid():
            form.save()
            messages.success(request, "El bono se ha añadido correctamente.")
            return redirect(reverse('client:voucher_client_list', args=[client.id]))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
    
    else:
        form = ClientVoucherForm(instance=voucher)

    return render(
        request,
        'voucher/client_edit.html',
        {'form': form, 'voucher': voucher, 'client': client})


def voucher_client_details(request, voucher_id):
    """
    View the details of a voucher associated with a client.

    Retrieves and displays the details of a specific voucher linked to a client.

    :param request: HttpRequest object with the request details.

    :param voucher_id: Integer ID of the voucher to view.

    :return: HttpResponse with the rendered 'voucher_client_details.html' template showing voucher details.
    """
    voucher = get_object_or_404(ClientVoucher, id=voucher_id)
    client = voucher.client
    return render(request, 'voucher/client_details.html', {'voucher': voucher, 'client': client})


def voucher_list(request, client_id):
    """
    Displays a list of vouchers associated with a client.

    This view retrieves all vouchers linked to a specific client and renders them in a template.

    :param request: HttpRequest object with the request details.

    :param client_id: Integer ID of the client whose vouchers to list.

    :return: HttpResponse with the rendered 'voucher_client_list.html' template displaying the vouchers.
    """
    client = get_object_or_404(Client, id=client_id)
    vouchers = ClientVoucher.objects.filter(client=client).select_related('voucher')
    
    return render(request, 'voucher/client_list.html', {'client': client, 'vouchers': vouchers})



    