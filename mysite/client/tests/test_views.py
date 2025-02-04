import io
import openpyxl
from openpyxl.workbook import Workbook
from xlwt import Workbook

from django.contrib.messages import get_messages
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse

from client.models import Client, ClientVoucher
from client.forms import ClientForm, ClientVoucherForm
from service.models import Voucher


# Test for client views
class AddClientViewTest(TestCase):
    """
    Test case for the AddClientView.

    This class contains tests to verify the functionality of the AddClientView, including successful client creation,
    handling of duplicate data, form validation, and GET request handling.
    """
    
    def setUp(self):
        """
        Set up the test environment.

        This method creates a Client object and prepares various data dictionaries
        for use in the tests.
        """
        self.client_obj = Client.objects.create(
            first_name="Juan",
            last_name="Perez",
            telephone_number="623456789",
            email="juan.perez@example.com"
        )
        self.valid_data = {
            'first_name': 'Ana',
            'last_name': 'Gómez',
            'telephone_number': '687654321',
            'email': 'ana.gomez@example.com',
        }
        self.duplicate_phone_data = {
            'first_name': 'Luis',
            'last_name': 'Martínez',
            'telephone_number': '623456789',  # Duplicate telephone number
            'email': 'luis.martinez@example.com',
        }
        self.duplicate_email_data = {
            'first_name': 'Carlos',
            'last_name': 'López',
            'telephone_number': '611222333',
            'email': 'juan.perez@example.com',  # Duplicate email
        }
    
    def test_add_client_success(self):
        """
        Test successful client creation.

        Verifies that a client is created successfully when valid data is submitted, and that the user is redirected to
        the client list page.
        """
        # Test that the client is created successfully
        response = self.client.post(reverse('client:client_add'), self.valid_data)
        
        self.assertEqual(response.status_code, 302)  # Successful redirection
        self.assertRedirects(response, reverse('client:client_list'))
        self.assertTrue(Client.objects.filter(email='ana.gomez@example.com').exists())
    
    def test_add_client_duplicate_phone(self):
        """
        Test client creation with duplicate phone number.

        Verifies that a client cannot be created with a phone number that already exists in the database, and that an
        appropriate error message is displayed.
        """
        response = self.client.post(reverse('client:client_add'), self.duplicate_phone_data)
        
        self.assertEqual(response.status_code, 200)  # Render the same page
        self.assertContains(response, "Ya existe un usuario con el número 623456789.")
        self.assertEqual(Client.objects.filter(telephone_number='623456789').count(), 1)
    
    def test_add_client_duplicate_email(self):
        """
        Test client creation with duplicate email.

        Verifies that a client cannot be created with an email that already exists in the database, and that an
        appropriate error message is displayed.
        """
        response = self.client.post(reverse('client:client_add'), self.duplicate_email_data)
        
        self.assertEqual(response.status_code, 200)  # Render the same page
        self.assertContains(response, "Ya existe un usuario con el email juan.perez@example.com.")
        self.assertEqual(Client.objects.filter(email='juan.perez@example.com').count(), 1)
    
    def test_add_client_invalid_form(self):
        """
        Test client creation with invalid form data.

        Verifies that a client cannot be created with invalid form data (empty fields), and that appropriate error
        messages are displayed.
        """
        invalid_data = {'first_name': '', 'last_name': '', 'telephone_number': '', 'email': ''}
        
        response = self.client.post(reverse('client:client_add'), invalid_data)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Este campo es obligatorio.")
        self.assertEqual(Client.objects.count(), 1)  # Only the initial client should exist
    
    def test_add_client_get_request(self):
        """
        Test GET request to add client page.

        Verifies that a GET request to the add client page returns the correct response,
        uses the correct template, and includes the ClientForm in the context.
        """
        response = self.client.get(reverse('client:client_add'))
        
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'client/add.html')
        self.assertIsInstance(response.context['form'], ClientForm)


class DeleteClientViewTest(TestCase):
    """
    Test case for the DeleteClientView.

    This class contains tests to verify the functionality of the DeleteClientView, including GET request handling and
    successful client deletion.
    """
    
    def setUp(self):
        """
        Set up the test environment.

        This method creates a Client object and sets up the URL for deletion tests.
        """
        self.client_obj = Client.objects.create(
            first_name="Juan",
            last_name="Perez",
            email="juan.perez@example.com",
            telephone_number="623456789"
        )
        self.url = reverse('client:client_delete', args=[self.client_obj.id])

    def test_delete_client_get(self):
        """
        Test GET request to delete client page.

        Verifies that a GET request to the delete client page returns the correct response, uses the correct template,
        and includes the expected confirmation message.
        """
        response = self.client.get(self.url)
        
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'client/delete.html')
        self.assertContains(response, '¿Estás segura de que deseas eliminar a <b>Juan Perez</b>?')
        
    def test_delete_client_post(self):
        """
        Test successful client deletion.

        Verifies that a POST request to the delete client page successfully deletes the client, redirects to the client
        list page, and displays a success message.
        """
        response = self.client.post(self.url)
        
        self.assertEqual(Client.objects.count(), 0)
        self.assertRedirects(response, reverse('client:client_list'))
        messages = list(get_messages(response.wsgi_request))
        self.assertEqual(str(messages[0]), 'Cliente eliminado correctamente.')


class EditClientViewTests(TestCase):
    """
    Test case for the EditClientView.

    This class contains tests to verify the functionality of the EditClientView, including GET request handling,
    successful client editing, handling of invalid data, and handling of non-existent clients.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method creates a Client object and sets up the URL for editing tests.
        """
        self.client_instance = Client.objects.create(
            first_name="Juan",
            last_name="Perez",
            email="juan.perez@example.com",
            telephone_number="623456789"
        )
        self.url = reverse('client:client_edit', args=[self.client_instance.id])

    def test_edit_client_get(self):
        """
        Test GET request to edit client page.

        Verifies that a GET request to the edit client page returns the correct response, uses the correct template,
        includes the ClientForm in the context, and provides the correct client instance.
        """
        response = self.client.get(self.url)
        
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'client/edit.html')
        self.assertIsInstance(response.context['form'], ClientForm)
        self.assertEqual(response.context['client'], self.client_instance)

    def test_edit_client_post_valid_data(self):
        """
        Test successful client editing with valid data.

        Verifies that a POST request with valid data successfully updates the client, redirects to the client list page,
        and correctly modifies the client data in the database.
        """
        valid_data = {
            'first_name': 'Carlos',
            'last_name': 'Lopez',
            'email': 'carlos.lopez@example.com',
            'telephone_number': '687654321'
        }
        response = self.client.post(self.url, data=valid_data)
        
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('client:client_list'))
        self.client_instance.refresh_from_db()
        self.assertEqual(self.client_instance.first_name, 'Carlos')
        self.assertEqual(self.client_instance.last_name, 'Lopez')
        self.assertEqual(self.client_instance.email, 'carlos.lopez@example.com')
        self.assertEqual(self.client_instance.telephone_number, '687654321')

    def test_edit_client_post_invalid_data(self):
        """
        Test client editing with invalid data.

        Verifies that a POST request with invalid data returns to the edit form, displays appropriate error messages,
        and does not modify the client data in the database.
        """
        invalid_data = {
            'first_name': '',  # Empty name isn't valid
            'last_name': 'Lopez',
            'email': 'invalid email',  # Invalid email
            'telephone_number': '687654321'
        }
        response = self.client.post(self.url, data=invalid_data)
        
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'client/edit.html')
        self.assertIsInstance(response.context['form'], ClientForm)
        self.assertFalse(response.context['form'].is_valid())
        self.client_instance.refresh_from_db()
        self.assertEqual(self.client_instance.first_name, 'Juan')
        self.assertEqual(self.client_instance.email, 'juan.perez@example.com')

    def test_edit_client_not_found(self):
        """
        Test editing a non-existent client.

        Verifies that attempting to edit a non-existent client returns a 404 error.
        """
        non_existent_url = reverse('client:client_edit', args=[9999])
        response = self.client.get(non_existent_url)
        
        self.assertEqual(response.status_code, 404)
        
        
class ExportClientsViewTests(TestCase):
    """
    Test case for the ExportClientsView.

    This class contains tests to verify the functionality of the ExportClientsView, including the correct generation and
    content of the exported Excel file.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method creates two Client objects and sets up the URL for export tests.
        """
        Client.objects.create(
            first_name="Juan",
            last_name="Perez",
            telephone_number="623456789",
            email="juan.perez@example.com"
        )
        Client.objects.create(
            first_name="Maria",
            last_name="Lopez",
            telephone_number="687654321",
            email="maria.lopez@example.com"
        )
        self.url = reverse('client:client_export')

    def test_export_clients(self):
        """
        Test the export of clients to an Excel file.

        This tests verifies that:
        1. The response has the correct status code and content type.
        2. The response includes the correct Content-Disposition header.
        3. The generated Excel file has the correct sheet name.
        4. The Excel file contains the correct headers.
        5. The Excel file contains the correct client data.
        """
        response = self.client.get(self.url)
        # Verify that content type is Excel
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment; filename=listado_clientes.xlsx', response['Content-Disposition'])
        # Read the Excel file of the response content
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        # Verify the sheet name
        self.assertEqual(ws.title, "Clientes")
        # Verify headers
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ['Nombre', 'Apellidos', 'Teléfono', 'email'])
        # Verify clients content
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[1], ('Juan', 'Perez', '623456789', 'juan.perez@example.com'))
        self.assertEqual(rows[2], ('Maria', 'Lopez', '687654321', 'maria.lopez@example.com'))
        
        
class ImportClientsViewTests(TestCase):
    """
    Test case for the ImportClientsView.
    
    This class contains tests to verify the functionality of the ImportClientsView, including the correct handling of
    XLSX and XLS file imports, as well as error handling for invalid file formats and form submissions.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method sets up the URL for import tests.
        """
        self.url = reverse('client:client_import')

    def create_xlsx_file(self, data):
        """
        Create a tests XLSX file with the given data.

        :param data: List of lists containing client data
        :return: BytesIO object containing the XLSX file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Apellidos', 'Teléfono', 'email'])  # Headers
        for row in data:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def create_xls_file(self, data):
        """
        Create a tests XLS file with the given data.

        :param data: List of lists containing client data
        :return: BytesIO object containing the XLS file
        """
        wb = Workbook()
        sheet = wb.add_sheet('Sheet 1')
        headers = ['Nombre', 'Apellidos', 'Teléfono', 'Email']
        for col, header in enumerate(headers):
            sheet.write(0, col, header)
        for row_idx, row in enumerate(data, start=1):
            for col_idx, value in enumerate(row):
                sheet.write(row_idx, col_idx, value)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def test_import_xlsx_file(self):
        """
        Test importing clients from an XLSX file.

        This tests verifies that:
        1. The view correctly processes an XLSX file.
        2. The correct number of clients are created in the database.
        3. The created clients have the correct data.
        """
        data = [
            ['Pedro', 'Ruiz', '623456789', 'pedro@example.com'],
            ['Laura', 'Gomez', '687654321', 'laura@example.com'],
        ]
        xlsx_file = self.create_xlsx_file(data)
        file = SimpleUploadedFile('clients.xlsx', xlsx_file.read(),
                                  content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        response = self.client.post(self.url, {'excel_file': file})
        
        self.assertRedirects(response, reverse('client:client_list'))
        self.assertEqual(Client.objects.count(), 2)
        self.assertTrue(Client.objects.filter(first_name='Pedro', email='pedro@example.com').exists())
        self.assertTrue(Client.objects.filter(first_name='Laura', email='laura@example.com').exists())

    def test_import_xls_file(self):
        """
        Test importing clients from an XLS file.

        This tests verifies that:
        1. The view correctly processes an XLS file.
        2. The correct number of clients are created in the database.
        """
        data = [
            ['Pedro', 'Ruiz', '623456789', 'pedro@example.com'],
            ['Laura', 'Gomez', '687654321', 'laura@example.com'],
        ]
        xls_file = self.create_xls_file(data)
        file = SimpleUploadedFile('clients.xls', xls_file.read(), content_type='application/vnd.ms-excel')

        response = self.client.post(self.url, {'excel_file': file})
        
        self.assertRedirects(response, reverse('client:client_list'))
        self.assertEqual(Client.objects.count(), 2)

    def test_invalid_file_format(self):
        """
        Test handling of an invalid file format.

        This tests verifies that:
        1. The view correctly handles an invalid file format.
        2. An appropriate error message is displayed.
        3. No clients are created in the database.
        """
        invalid_file = SimpleUploadedFile('clients.txt', b'Invalid content', content_type='text/plain')
        response = self.client.post(self.url, {'excel_file': invalid_file})
        
        self.assertContains(response, "Por favor, sube un archivo válido.", status_code=200)
        self.assertEqual(Client.objects.count(), 0)

    def test_invalid_form_submission(self):
        """
        Test handling of an invalid form submission.

        This tests verifies that:
        1. The view correctly handles an invalid form submission (no file uploaded).
        2. An appropriate error message is displayed.
        3. No clients are created in the database.
        """
        response = self.client.post(self.url, {})
        
        self.assertContains(response, "Por favor, sube un archivo válido.", status_code=200)
        self.assertEqual(Client.objects.count(), 0)


class ClientListViewTests(TestCase):
    """
    Test case for the ClientListView.

    This class contains tests to verify the functionality of the ClientListView, including pagination, template usage,
    and handling of various page scenarios.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method creates 25 tests clients for pagination testing.
        """
        for i in range(25):
            Client.objects.create(
                first_name=f'Client{i}',
                last_name=f'Last{i}',
                telephone_number=f'666-222-333{i}',
                email=f'client{i}@example.com'
            )
    
    def test_client_list_view_status_code(self):
        """
        Test that the ClientListView returns a 200 status code.
        """
        url = reverse('client:client_list')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
    
    def test_client_list_view_template(self):
        """
        Test that the ClientListView uses the correct template.
        """
        url = reverse('client:client_list')
        response = self.client.get(url)
        
        self.assertTemplateUsed(response, 'client/list.html')
    
    def test_pagination(self):
        """
        Test that pagination works correctly for the first page.

        Verifies that the first page contains 20 clients.
        """
        url = reverse('client:client_list')
        response = self.client.get(url)
        
        self.assertEqual(len(response.context['clients']), 20)
    
    def test_pagination_second_page(self):
        """
        Test that pagination works correctly for the second page.

        Verifies that the second page contains 5 clients.
        """
        url = reverse('client:client_list') + '?page=2'
        response = self.client.get(url)
        
        self.assertEqual(len(response.context['clients']), 5)
    
    def test_invalid_page_number(self):
        """
        Test handling of an invalid page number.

        Verifies that an invalid page number defaults to the first page.
        """
        url = reverse('client:client_list') + '?page=abc'
        response = self.client.get(url)
        
        # Check that the first page is shown (20 clients)
        self.assertEqual(len(response.context['clients']), 20)
    
    def test_page_out_of_range(self):
        """
        Test handling of an out-of-range page number.

        Verifies that an out-of-range page number returns the last page.
        """
        url = reverse('client:client_list') + '?page=999'
        response = self.client.get(url)
        # Check that the last page is shown (5 clients)
        self.assertEqual(len(response.context['clients']), 5)


class SearchClientViewTests(TestCase):
    """
    Test case for the SearchClientView.

    This class contains tests to verify the functionality of the SearchClientView, including various search scenarios
    and edge cases.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method creates three tests clients for search testing.
        """
        Client.objects.create(first_name="Juan", last_name="Perez", telephone_number="23456789",
                              email="juan.perez@example.com")
        Client.objects.create(first_name="Maria", last_name="Leon", telephone_number="687654321",
                              email="maria.leon@example.com")
        Client.objects.create(first_name="Alicia", last_name="Martin", telephone_number="689541236",
                              email="alicia.martin@example.com")
    
    def test_search_by_first_name(self):
        """
        Test searching clients by first name.

        Verifies that the search returns the correct client when searching by first name.
        """
        url = reverse('client:client_search') + '?query=Juan'
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Juan Perez")
        self.assertNotContains(response, "Maria Leon")
    
    def test_search_by_last_name(self):
        """
        Test searching clients by last name.

        Verifies that the search returns the correct client when searching by last name.
        """
        url = reverse('client:client_search') + '?query=Leon'
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Maria Leon")
        self.assertNotContains(response, "Juan Perez")
    
    def test_search_by_full_name(self):
        """
        Test searching clients by full name.

        Verifies that the search returns the correct client when searching by full name.
        """
        url = reverse('client:client_search') + '?query=Juan Perez'
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Juan Perez")
        self.assertNotContains(response, "Maria Leon")
    
    def test_search_by_partial_name(self):
        """
        Test searching clients by partial name.

        Verifies that the search returns the correct client when searching by a partial name.
        """
        url = reverse('client:client_search') + '?query=Jua'
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Juan Perez")
        self.assertNotContains(response, "Alicia Martin")
    
    def test_search_by_telephone_number(self):
        """
        Test searching clients by telephone number.

        Verifies that the search returns the correct client when searching by telephone number.
        """
        url = reverse('client:client_search') + '?query=689541236'
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Alicia Martin")
        self.assertNotContains(response, "Juan Perez")
    
    def test_search_no_results(self):
        """
        Test searching with a query that returns no results.

        Verifies that the search displays the correct message when no clients are found.
        """
        url = reverse('client:client_search') + '?query=Nonexistent'
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No se encontró ningún cliente que coincida con los criterios de búsqueda.")
    
    def test_empty_query(self):
        """
        Test searching with an empty query.

        Verifies that the search handles an empty query correctly.
        """
        url = reverse('client:client_search') + '?query='
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No se encontró ningún cliente que coincida con los criterios de búsqueda.")
        self.assertEqual(response.context['clients'], [])  # Verify that there aren't clients
    
    def test_query_with_extra_spaces(self):
        """
        Test searching with a query containing extra spaces.
    
        Verifies that the search handles queries with extra spaces correctly.
        """
        url = reverse('client:client_search') + '?query=  Juan  '
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Juan Perez")
        self.assertNotContains(response, "Maria Leon")
        
        
class ViewDetailClientTests(TestCase):
    """
    Test case for the ViewDetailClient view.

    This class contains tests to verify the functionality of the ViewDetailClient view, including viewing existing
    clients, handling non-existent clients, and checking the rendered content.
    """
    
    def setUp(self):
        """
       Set up the tests environment.

       This method creates a tests client instance to be used in the tests.
       """
        self.client_instance = Client.objects.create(
            first_name="Juan",
            last_name="Perez",
            telephone_number="623456789",
            email="juan.perez@example.com"
        )

    def test_view_existing_client(self):
        """
        Test viewing an existing client.

        This tests verifies that:
        1. The view returns a 200 status code for an existing client.
        2. The correct template is used.
        3. The response contains the client's details.
        """
        url = reverse('client:client_view', args=[self.client_instance.id])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'client/view.html')
        self.assertContains(response, self.client_instance.first_name)
        self.assertContains(response, self.client_instance.last_name)
        self.assertContains(response, self.client_instance.telephone_number)
        self.assertContains(response, self.client_instance.email)

    def test_view_nonexistent_client(self):
        """
        Test viewing a non-existent client.

        This tests verifies that the view returns a 404 status code when attempting to view a client that doesn't exist.
        """
        url = reverse('client:client_view', args=[9999])  # Non-existent id
        response = self.client.get(url)

        self.assertEqual(response.status_code, 404)

    def test_view_content_rendered(self):
        """
        Test the rendered content of the view.

        This tests verifies that the client's details are correctly rendered in the response content.
        """
        url = reverse('client:client_view', args=[self.client_instance.id])
        response = self.client.get(url)
        
        self.assertIn(self.client_instance.first_name, response.content.decode())
        self.assertIn(self.client_instance.last_name, response.content.decode())
        self.assertIn(self.client_instance.telephone_number, response.content.decode())
        self.assertIn(self.client_instance.email, response.content.decode())
        
        
# Test for client voucher views
class AssignVoucherViewTests(TestCase):
    """
    Test case for the AssignVoucherView.

    This class contains tests to verify the functionality of the AssignVoucherView, including accessing the view,
    assigning vouchers, and rendering the correct template.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method creates tests instances of Client, Voucher, and ClientVoucher to be used in the tests.
        """
        self.client_instance = Client.objects.create(
            first_name="Maria",
            last_name="Leon",
            telephone_number="687654321",
            email="maria.leon@example.com"
        )
        self.voucher_instance = Voucher.objects.create(
            name="Masajes",
            total_sessions=10,
            discounted_price=350
        )
        self.client_voucher_instance = ClientVoucher.objects.create(
            client=self.client_instance,
            voucher=self.voucher_instance,
            purchase_date="2025-01-01",
            expiration_date="2025-01-31",
            sessions_remaining=10,
            is_active=True
        )

    def test_access_view_with_existing_client(self):
        """
        Test accessing the view with an existing client.

        This tests verifies that:
        1. The view returns a 200 status code for an existing client.
        2. The correct template is used.
        3. The response contains the client's name and voucher information.
        """
        url = reverse('client:voucher_assign', args=[self.client_instance.id])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'voucher/assign.html')
        self.assertContains(response, self.client_instance.first_name)
        self.assertContains(response, self.client_voucher_instance.voucher)

    def test_access_view_with_nonexistent_client(self):
        """
        Test accessing the view with a non-existent client.

        This tests verifies that the view returns a 404 status code when attempting to access it with a non-existent
        client ID.
        """
        url = reverse('client:voucher_assign', args=[9999])  # Non-existent id
        response = self.client.get(url)

        self.assertEqual(response.status_code, 404)

    def test_assign_voucher(self):
        """
        Test assigning a voucher to a client.

        This tests verifies that:
        1. The view correctly handles a POST request to assign a voucher.
        2. The ClientVoucher object is created in the database.
        """
        url = reverse('client:voucher_assign', args=[self.client_instance.id])
        data = {
            'client_voucher': self.client_voucher_instance.id,
        }
        response = self.client.post(url, data)

        self.assertEqual(response.status_code, 200)  # Redirection after assigning
        self.assertTrue(
            ClientVoucher.objects.filter(client=self.client_instance, voucher=self.voucher_instance).exists()
        )
        
    def test_template_renders_client_and_vouchers(self):
        """
        Test that the template correctly renders client and voucher information.

        This tests verifies that the response contains the client's name and the assigned voucher information.
        """
        url = reverse('client:voucher_assign', args=[self.client_instance.id])
        response = self.client.get(url)

        self.assertContains(response, self.client_instance.first_name)
        self.assertContains(response, self.client_voucher_instance.voucher)


class DeleteClientVoucherTests(TestCase):
    """
    Test case for the DeleteClientVoucher view.

    This class contains tests to verify the functionality of the DeleteClientVoucher view, including the successful
    deletion of a ClientVoucher instance.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method creates tests instances of Client, Voucher, and ClientVoucher to be used in the tests.
        """
        self.client_instance = Client.objects.create(
            first_name="Maria",
            last_name="Leon",
            telephone_number="623456789",
            email="maria.leon@example.com"
        )
        
        self.voucher_instance = Voucher.objects.create(
            name="Voucher de prueba",
            total_sessions=10,
            discounted_price=250
        )
        
        self.client_voucher_instance = ClientVoucher.objects.create(
            client=self.client_instance,
            voucher=self.voucher_instance,
            purchase_date="2025-01-01",
            expiration_date="2025-12-31",
            sessions_remaining=10
        )
    
    def test_delete_client_voucher(self):
        """
        Test the deletion of a ClientVoucher instance.

        This tests verifies that:
        1. The ClientVoucher instance is successfully deleted from the database.
        2. The view redirects to the correct page after deletion.
        3. A success message is displayed after the deletion.
        """
        url = reverse('client:voucher_client_delete', args=[self.client_voucher_instance.id])
        response = self.client.post(url)
        
        self.assertFalse(ClientVoucher.objects.filter(id=self.client_voucher_instance.id).exists())
        self.assertRedirects(response, reverse('client:voucher_client_list', args=[self.client_instance.id]))
        messages = list(get_messages(response.wsgi_request))
        self.assertEqual(str(messages[0]), "Bono eliminado correctamente.")


class EditClientVoucherTests(TestCase):
    """
    Test case for the EditClientVoucher view.

    This class contains tests to verify the functionality of the EditClientVoucher view, including GET and POST requests
    with valid and invalid data.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method creates tests instances of Client, Voucher, and ClientVoucher to be used in the tests.
        """
        self.client_instance = Client.objects.create(
            first_name="Maria Leon",
            last_name="Masajes",
            telephone_number="623456789",
            email="maria@example.com"
        )
        self.voucher_instance = Voucher.objects.create(
            name="Voucher de prueba",
            total_sessions=10
        )
        self.client_voucher_instance = ClientVoucher.objects.create(
            client=self.client_instance,
            voucher=self.voucher_instance,
            purchase_date="2025-01-01",
            expiration_date="2025-12-31",
            sessions_remaining=10
        )
    
    def test_edit_client_voucher_get(self):
        """
        Test the GET request for editing a ClientVoucher.

        This tests verifies that:
        1. The view returns a 200 status code.
        2. The correct form is present in the context.
        3. The form instance matches the ClientVoucher being edited.
        """
        url = reverse('client:voucher_client_edit', args=[self.client_voucher_instance.id])
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertIsInstance(response.context['form'], ClientVoucherForm)
        self.assertEqual(response.context['form'].instance, self.client_voucher_instance)
    
    def test_edit_client_voucher_post_valid_data(self):
        """
        Test the POST request for editing a ClientVoucher with valid data.

        This tests verifies that:
        1. The view redirects to the correct page after successful edit.
        2. The ClientVoucher data is updated correctly in the database.
        """
        url = reverse('client:voucher_client_edit', args=[self.client_voucher_instance.id])
        updated_data = {
            'client': self.client_instance.id,
            'voucher': self.voucher_instance.id,
            'purchase_date': '2025-03-01',
            'expiration_date': '2025-03-31',
        }
        
        response = self.client.post(url, updated_data)

        self.assertRedirects(response, reverse('client:voucher_client_list', args=[self.client_instance.id]))
        self.client_voucher_instance.refresh_from_db()
        self.assertEqual(self.client_voucher_instance.purchase_date.strftime('%Y-%m-%d'), '2025-03-01')
    
    def test_edit_client_voucher_post_invalid_data(self):
        """
        Test the POST request for editing a ClientVoucher with invalid data.

        This tests verifies that:
        1. The view returns a 200 status code when invalid data is submitted.
        2. The form contains appropriate error messages.
        3. The ClientVoucher data is not updated in the database.
        """
        url = reverse('client:voucher_client_edit', args=[self.client_voucher_instance.id])
        invalid_data = {
            'client': self.client_instance.id,
            'voucher': self.voucher_instance.id,
            'purchase_date': '2025-12-31',
            'expiration_date': '2025-11-01',
        }
        
        response = self.client.post(url, invalid_data)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'La fecha de expiración debe ser posterior a la fecha de compra.')
        self.client_voucher_instance.refresh_from_db()
        self.assertNotEqual(self.client_voucher_instance.purchase_date.strftime('%Y-%m-%d'), '2025-12-31')
        self.assertNotEqual(self.client_voucher_instance.expiration_date.strftime('%Y-%m-%d'), '2025-11-01')


class VoucherClientDetailsViewTests(TestCase):
    """
    Test case for the VoucherClientDetailsView.
    
    This class contains tests to verify the functionality of the VoucherClientDetailsView, including URL accessibility,
    template usage, context data, and displayed content.
    """
    
    def setUp(self):
        """
        Set up the tests environment.

        This method creates tests instances of Client, Voucher, and ClientVoucher to be used in the tests.
        """
        self.client_instance = Client.objects.create(
            first_name='Pablo',
            last_name='Mendez',
            telephone_number="623456789",
            email="pablo.m@example.com")
        
        self.voucher_instance = Voucher.objects.create(
            name="Masajes",
            total_sessions=10
        )

        self.client_voucher = ClientVoucher.objects.create(
            client=self.client_instance,
            voucher=self.voucher_instance,
            purchase_date="2025-01-20",
            expiration_date="2025-02-19"
        )
    
    def test_view_url_exists_at_desired_location(self):
        """
        Test that the view URL exists at the desired location.
        """
        response = self.client.get(f'/client/client_details/{self.client_voucher.id}/')
        
        self.assertEqual(response.status_code, 200)
    
    def test_view_url_accessible_by_name(self):
        """
        Test that the view URL is accessible by name.
        """
        response = self.client.get(reverse('client:voucher_client_details', args=[self.client_voucher.id]))
        
        self.assertEqual(response.status_code, 200)
    
    def test_view_uses_correct_template(self):
        """
        Test that the view uses the correct template.
        """
        response = self.client.get(reverse('client:voucher_client_details', args=[self.client_voucher.id]))
        
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'voucher/client_details.html')
    
    def test_context_data(self):
        """
        Test that the view provides the correct context data.
        """
        response = self.client.get(reverse('client:voucher_client_details', args=[self.client_voucher.id]))
        
        self.assertEqual(response.status_code, 200)
        self.assertIn('voucher', response.context)
        self.assertIn('client', response.context)
        self.assertEqual(response.context['voucher'], self.client_voucher)
        self.assertEqual(response.context['client'], self.client_instance)
    
    def test_voucher_details_displayed(self):
        """
        Test that the voucher details are correctly displayed in the response.
        """
        
        response = self.client.get(reverse('client:voucher_client_details', args=[self.client_voucher.id]))
        
        self.assertContains(response, "Masajes")
        self.assertContains(response, "Pablo")
        self.assertContains(response, "Mendez")
        self.assertContains(response, "20 de enero de 2025")
        self.assertContains(response, "19 de febrero de 2025")


class VoucherListViewTests(TestCase):
    """
    Test case for the VoucherListView.
    
    This class contains tests to verify the functionality of the VoucherListView, including the correct display of
    vouchers for a specific client.
    """
    def setUp(self):
        """
        Set up the tests environment.

        This method creates tests instances of Client, Voucher, and ClientVoucher to be used in the tests.
        """
        self.client_instance = Client.objects.create(
            first_name="Juan",
            last_name="Perez",
            telephone_number="623456789",
            email="juan.perez@example.com"
        )

        self.voucher_1 = Voucher.objects.create(name="Voucher 1", total_sessions=10)
        self.voucher_2 = Voucher.objects.create(name="Voucher 2", total_sessions=10)
        
        self.client_voucher_1 = ClientVoucher.objects.create(
            client=self.client_instance,
            voucher=self.voucher_1,
            purchase_date="2025-01-01",
            expiration_date="2025-01-31"
        )
        self.client_voucher_2 = ClientVoucher.objects.create(
            client=self.client_instance,
            voucher=self.voucher_2,
            purchase_date="2025-03-01",
            expiration_date="2025-03-31"
        )
    
    def test_voucher_list_view(self):
        """
        Test the VoucherListView.

        This tests verifies that:
        1. The view returns a 200 status code.
        2. The correct template is used.
        3. The context contains the correct vouchers and client.
        4. The response contains the correct voucher names.
        """
        url = reverse('client:voucher_client_list', args=[self.client_instance.id])
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'voucher/client_list.html')
        
        vouchers = response.context['vouchers']
        self.assertIn(self.client_voucher_1, vouchers)
        self.assertIn(self.client_voucher_2, vouchers)
        
        client = response.context['client']
        self.assertEqual(client, self.client_instance)
        self.assertContains(response, "Voucher 1")
        self.assertContains(response, "Voucher 2")
        
        
